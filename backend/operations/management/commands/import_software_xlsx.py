from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from openpyxl import load_workbook

from operations.models import Brand, Category, Product


class _DryRunRollback(Exception):
  pass


def _clean(value):
  if value is None:
    return ''
  return ' '.join(str(value).strip().split())


class Command(BaseCommand):
  help = 'Import BRAND/CATEGORY/PRODUCT rows from SOFTWARE.xlsx into operations master data.'

  def add_arguments(self, parser):
    parser.add_argument(
      '--path',
      default=r'C:\Users\usman\Downloads\SOFTWARE.xlsx',
      help='Absolute path to the XLSX file.',
    )
    parser.add_argument(
      '--sheet',
      default='Sheet1',
      help='Worksheet name to import from.',
    )
    parser.add_argument(
      '--dry-run',
      action='store_true',
      help='Parse and validate rows without committing database changes.',
    )

  def handle(self, *args, **options):
    xlsx_path = Path(options['path'])
    sheet_name = options['sheet']
    dry_run = options['dry_run']

    if not xlsx_path.exists():
      raise CommandError(f'File not found: {xlsx_path}')

    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
      raise CommandError(f'Sheet "{sheet_name}" not found. Available sheets: {workbook.sheetnames}')

    ws = workbook[sheet_name]
    header = [str(col).strip().upper() if col is not None else '' for col in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    if header[:3] != ['BRAND', 'CATEGORY', 'PRODUCT']:
      raise CommandError(f'Expected first 3 headers to be BRAND, CATEGORY, PRODUCT. Got: {header[:3]}')

    created_brands = 0
    created_categories = 0
    created_products = 0
    updated_products = 0
    skipped_rows = 0
    processed_rows = 0

    current_brand = ''
    current_category = ''

    try:
      with transaction.atomic():
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
          brand_raw = _clean(row[0] if len(row) > 0 else '')
          category_raw = _clean(row[1] if len(row) > 1 else '')
          product_raw = _clean(row[2] if len(row) > 2 else '')

          if brand_raw:
            current_brand = brand_raw
          if category_raw:
            current_category = category_raw

          if not product_raw:
            skipped_rows += 1
            continue

          if not current_brand or not current_category:
            skipped_rows += 1
            self.stdout.write(self.style.WARNING(
              f'Skipped row {row_idx}: missing brand/category context for product "{product_raw}"'
            ))
            continue

          processed_rows += 1

          brand, brand_created = Brand.objects.get_or_create(
            name=current_brand,
            defaults={'status': True},
          )
          if brand_created:
            created_brands += 1

          category, category_created = Category.objects.get_or_create(
            name=current_category,
            brand=brand,
            defaults={'status': True},
          )
          if category_created:
            created_categories += 1

          product, product_created = Product.objects.get_or_create(
            name=product_raw,
            category=category,
            defaults={'brand': brand, 'status': True},
          )
          if product_created:
            created_products += 1
          else:
            # Keep brand linkage consistent in case existing rows are partial.
            dirty = False
            if product.brand_id != brand.id:
              product.brand = brand
              dirty = True
            if product.status is False:
              product.status = True
              dirty = True
            if dirty:
              product.save(update_fields=['brand', 'status', 'updated_at'])
              updated_products += 1

        if dry_run:
          raise _DryRunRollback()
    except _DryRunRollback:
      self.stdout.write(self.style.WARNING('Dry run complete. No database changes were committed.'))

    self.stdout.write('')
    self.stdout.write(self.style.SUCCESS('Import summary'))
    self.stdout.write(f'File: {xlsx_path}')
    self.stdout.write(f'Sheet: {sheet_name}')
    self.stdout.write(f'Rows processed: {processed_rows}')
    self.stdout.write(f'Rows skipped: {skipped_rows}')
    self.stdout.write(f'Brands created: {created_brands}')
    self.stdout.write(f'Categories created: {created_categories}')
    self.stdout.write(f'Products created: {created_products}')
    self.stdout.write(f'Products updated: {updated_products}')
